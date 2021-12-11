from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from datetime import date, time, timedelta
import time
import unittest
import HtmlTestRunner
import openpyxl

print('OneShield_UnderLyingPolicies Script')
try:
    class FL06(unittest.TestCase):
        from CommonFunctions import BasicFunctions as BF, OneShieldFunctions as OSF
        import AccessorRepository as AR

        uLayers = myData['Option1']

        myDriver = myDriver
        ScreenShotsDir = ScreenShotsDir
        testDataLoc = testDataLoc

        def test_FL06_01_GoToULPolicies(self):
            try:
                # Go to Under Lying Policies Tab
                self.myDriver.find_element(By.CSS_SELECTOR, "#ctl00__MainContent_mnuDataTabs > .rmRootGroup > .rmItem:nth-child(5) .rmText").click()
                time.sleep(3)

            except Exception as e51:
                print(self.BF.fncCaptureScreenshot(self.myDriver, self.ScreenShotsDir, 'GoToULPolicies'))
                self.fail(e51)

        def test_FL06_02_DeleteOldPolicies(self):
            try:
                # Delete any existing policies
                idbtnULPDeleteAll = 'ctl00__MainContent_ucDataCtrl_btnDeleteAll_input'
                idbtnULPConfirmYes = 'ctl00__MainContent_confirmWindow_C_btnYes'
                # If there are any existing policies, delete them all
                WebDriverWait(self.myDriver, 10).until(EC.element_to_be_clickable((By.ID, idbtnULPDeleteAll))).click()
                time.sleep(3)

                # Click Yes on the Confirmation Dialog pop up
                WebDriverWait(self.myDriver, 10).until(EC.element_to_be_clickable((By.ID, idbtnULPConfirmYes))).click()
                time.sleep(3)

            except Exception as e52:
                # Do nothing, do not fail the script
                y=1


        def test_FL06_03_AddLayers(self):
            try:

                wb = openpyxl.load_workbook(self.testDataLoc)
                sheet = wb['UnderlyingPolicies']

                for i in range(1, int(self.uLayers)+1):
                    uLayerNumber = sheet.cell(row=i, column=1).value
                    uPolicyType = sheet.cell(row=i, column=2).value
                    uCarrier = sheet.cell(row=i, column=3).value
                    uULCoverageCode = sheet.cell(row=i, column=4).value
                    uULCoverageType = sheet.cell(row=i, column=5).value
                    uPremium = sheet.cell(row=i, column=6).value
                    uULRatedSUPolicy = sheet.cell(row=i, column=7).value
                    uRCCLimit = sheet.cell(row=i, column=8).value
                    uRetention = sheet.cell(row=i, column=9).value

                    self.OSF.fncAddUnderLyingPolicyCoverage(self.myDriver, uLayerNumber, uPolicyType, uCarrier, uULCoverageCode, uULCoverageType, uPremium, uULRatedSUPolicy, uRCCLimit, uRetention)

                # self.OSF.fncAddUnderLyingPolicyCoverage(self.myDriver, self.uLayerNumber1, self.uCarrier1,
                #                                         self.uULCoverageCode1, self.uULCoverageType1, self.uPremium1,
                #                                         self.uULRatedSUPolicy1, self.uRCCLimit1, self.uRetention1)
                #
                # if self.uLayerNumber2 != "":
                #     self.OSF.fncAddUnderLyingPolicyCoverage(self.myDriver, self.uLayerNumber2, self.uCarrier2,
                #                                         self.uULCoverageCode2, self.uULCoverageType2, self.uPremium2,
                #                                         self.uULRatedSUPolicy2, self.uRCCLimit2, self.uRetention2)
                #
                # if self.uLayerNumber3 != "":
                #     self.OSF.fncAddUnderLyingPolicyCoverage(self.myDriver, self.uLayerNumber3, self.uCarrier3,
                #                                         self.uULCoverageCode3, self.uULCoverageType3, self.uPremium3,
                #                                         self.uULRatedSUPolicy3, self.uRCCLimit3, self.uRetention3)
                #
                # if self.uLayerNumber4 != "":
                #     self.OSF.fncAddUnderLyingPolicyCoverage(self.myDriver, self.uLayerNumber4, self.uCarrier4,
                #                                         self.uULCoverageCode4, self.uULCoverageType4, self.uPremium4,
                #                                         self.uULRatedSUPolicy4, self.uRCCLimit4, self.uRetention4)
                #
                # if self.uLayerNumber5 != "":
                #     self.OSF.fncAddUnderLyingPolicyCoverage(self.myDriver, self.uLayerNumber5, self.uCarrier5,
                #                                         self.uULCoverageCode5, self.uULCoverageType5, self.uPremium5,
                #                                         self.uULRatedSUPolicy5, self.uRCCLimit5, self.uRetention5)
                # Go to Next page
                self.myDriver.find_element(By.ID, self.AR.idbtnOneShieldNext).click()
                time.sleep(2)


            except Exception as e53:
                print(self.BF.fncCaptureScreenshot(self.myDriver, self.ScreenShotsDir, 'AddLayers'))
                self.fail(e53)

except Exception as e6:
    print('OneShield_UnderLyingPolicies Script did not run successfully')
    print(e6)

