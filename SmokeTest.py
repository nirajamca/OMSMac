import os
import sys
from datetime import datetime
import openpyxl
from CommonFunctions import BasicFunctions as BF
import unittest
import HtmlTestRunner

# Open the test fate spreadsheet from workbook
wb = openpyxl.load_workbook(testDataLoc, read_only=True)
sheet = wb.active
sheet = wb[testDataSheet]

myDriver = BF.fncGetMyDriver()

# Skipping the first row[header], capture each test case and whether to execute each test case
for i in range(2, sheet.max_row+1):
    testCase = sheet.cell(row=i, column=1).value
    toExecute = sheet.cell(row=i, column=2).value
    if(toExecute == "Y"):

        # Get data from spreadsheet
        myData = BF.fncGetValues(testDataLoc, testDataSheet, testCase)

        # execute each of the test cases
        exec(open(testCase).read())


wb.close()

if __name__ == '__main__':
    unittest.main(testRunner=HtmlTestRunner.HTMLTestRunner(output='reports', combine_reports="True"))
