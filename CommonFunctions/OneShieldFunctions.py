
# *******************************************************************************************************************************
def fncAddUnderLyingPolicyCoverage(myDriver, uLayerNumber, uPolicyType, uCarrier, uULCoverageCode, uULCoverageType, uPremium, uULRatedSUPolicy, uRCCLimit, uRetention):
    import AccessorRepository as AR
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.common.action_chains import ActionChains
    from selenium.webdriver.common.keys import Keys
    import time
    from CommonFunctions import BasicFunctions as BF

    # Click Add button to add policies
    WebDriverWait(myDriver, 10).until(EC.element_to_be_clickable((By.ID, AR.idbtnULPolicyAdd))).click()
    time.sleep(3)
    
    # Layer
    BF.fncEnterTextboxValue(myDriver, AR.idtxtULPoliciesLayerNumber, int(uLayerNumber))

    # Policy Type
    # BF.fncSelectElementFromDropdown(myDriver, AR.idddULPPolicyType, uPolicyType)

    # Carrier
    myDriver.find_element(By.ID, AR.idtxtULCarrier).click()
    myDriver.find_element(By.ID, AR.idtxtULCarrier).clear()
    myDriver.find_element(By.ID, AR.idtxtULCarrier).send_keys(str(uCarrier))
    time.sleep(2)
    myDriver.find_element(By.ID, AR.idtxtULCarrier).send_keys(Keys.DOWN)
    time.sleep(2)
    myDriver.find_element(By.ID, AR.idtxtULCarrier).send_keys(Keys.RETURN)
    time.sleep(2)

    # Under Lying Coverage Code
    BF.fncSelectElementFromDropdown(myDriver, AR.idddULCoverageCode, uULCoverageCode)

    # Coverage Type
    BF.fncSelectElementFromDropdown(myDriver, AR.idddULCoverageType, uULCoverageType)

    # Premium
    BF.fncEnterTextboxValue(myDriver, AR.idtxtULPoliciesPremium, uPremium)

    # Rated Scheduled Underlying Policy
    BF.fncSelectElementFromDropdown(myDriver, AR.idddULRatedSUPolicy, uULRatedSUPolicy)

    # Coverage table
    # General Aggregate Checkbox
    WebDriverWait(myDriver, 10).until(EC.element_to_be_clickable((By.ID, AR.idchkULRiskCoverageControl))).click()
    
    # Limit
    BF.fncEnterTextboxValue(myDriver, AR.idtxtULRCCLimit, uRCCLimit)

    # Deductible/Retention
    BF.fncEnterTextboxValue(myDriver, AR.idtxtULRCCRetension, uRetention)

    # Save the coverage
    WebDriverWait(myDriver, 10).until(EC.element_to_be_clickable((By.ID, AR.idbtnULRCCSave))).click()
    time.sleep(2)

# *******************************************************************************************************************************
def fncGetStateRangeAllowed(uRangeState):
    import openpyxl
    loc = 'Submissions\State List of Range states vs Non Range states.xlsx'
    wb = openpyxl.load_workbook(loc)
    sheet = wb['Sheet1']


    for i in range(4, 54):
        if (sheet.cell(row=i, column=1).value == uRangeState):
            return sheet.cell(i, 2).value

# *******************************************************************************************************************************
def fncCheckIfFileExists(uFilename):
    import os
    from openpyxl import Workbook
    if os.path.isfile(uFilename) and os.access(uFilename, os.R_OK):
        print("File exists and is readable, can proceed adding values")
    else:
        print("Either the file is missing or not readable, creating now")
        wb = Workbook()
        sheet = wb.active
        sheet.title = 'CaptureValuesFromOMS'
        sheet['A1'] = 'Item'
        sheet['B1'] = 'Value'

        wb.save(uFilename)
        wb.close()

def fncAddDataToGLBaseFile(uFilename, uData, uValue):
    import openpyxl

    fncCheckIfFileExists(uFilename)

    wb = openpyxl.open(uFilename)
    sheet = wb.active

    xrow = sheet.max_row + 1

    sheet.cell(row=xrow, column=1).value = uData
    sheet.cell(row=xrow, column=2).value = uValue

    wb.save(uFilename)
    wb.close()