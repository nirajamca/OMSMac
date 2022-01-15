def fncGetMyDriver():
    from selenium import webdriver
    import chromedriver_autoinstaller

    # Create and object for Chrome WebDriver
    # myDriver = webdriver.Chrome(executable_path='C:\Drivers\chromedriver.exe')
    chromedriver_autoinstaller.install()
    myDriver = webdriver.Chrome()
    myDriver.implicitly_wait(2)
    myDriver.maximize_window()

    return myDriver

# *********************************************************************************************************************
# Name: fncCreateDirectory
# Purpose: tries to create a directly specified in dirName, only if it is not existing already
#     param 1: dirName
# Usage: fncCreateDirectory(dirName)
# Example: fncCreateDirectory("C:\\TestResults")
# *********************************************************************************************************************
def fncCreateDirectory(dirName):
    import os
    try:
        # Create target Directory
        os.mkdir(dirName)
    except FileExistsError:
        i = 1


# ===============================================================================
def fncCaptureScreenshot(myDriver, testResultsDir, uFilename):
    from datetime import datetime

    # Create directory. Only if it not already existing
    fncCreateDirectory(testResultsDir)

    # Capture current date/time to log in the test results
    currentDateTime = datetime.now().strftime("_%m-%d-%Y_%H-%M-%S")
    imgFilename = testResultsDir + "\\" + uFilename + "" + currentDateTime + ".png"
    myDriver.save_screenshot(imgFilename)

    return imgFilename


# # ===============================================================================
# def fncGetValues(loc, sheetName, testCase):
#     # import openpyxl module to use excel operations
#     import openpyxl
#
#     # Create workbook object and capture the active sheet
#     wb = openpyxl.load_workbook(loc)
#     sheet = wb.active
#
#     # Designate the actual sheetname
#     sheet = wb[sheetName]
#
#     # Leaving the header row, go through each line to find a match with test case you want to execute
#     for i in range(2, sheet.max_row + 1):
#         if (sheet.cell(row=i, column=1).value == testCase):
#
#             # if the matching test case is found, create a dictionary with data in second column
#             getValues = {sheet.cell(row=1, column=2).value: sheet.cell(row=i, column=2).value}
#
#             # Once the data dictionary is added, keep on adding rest of the items in that row till the end of columns
#             for j in range(3, sheet.max_column + 1):
#                 getValues.update({sheet.cell(row=1, column=j).value: sheet.cell(row=i, column=j).value})
#
#             break
#
#     # Return the dictionary
#     return getValues

# ===============================================================================
# def fncGetValues(loc, sheetName, testCase):
#     # import openpyxl module to use excel operations
#     import xlrd
#
#     # Create workbook object and capture the active sheet
#     wb = xlrd.open_workbook(loc)
#
#     # Designate the actual sheetname
#     sheet = wb.sheet_by_name(sheetName)
#
#     # Leaving the header row, go through each line to find a match with test case you want to execute
#     for i in range(1, sheet.utter_max_rows):
#         if (sheet.cell(rowx=i, colx=0).value == testCase):
#
#             # if the matching test case is found, create a dictionary with data in second column
#             getValues = {sheet.cell(rowx=0, colx=1).value: sheet.cell(rowx=i, colx=1).value}
#
#             # Once the data dictionary is added, keep on adding rest of the items in that row till the end of columns
#             for j in range(2, sheet.ncols):
#                 getValues.update({sheet.cell(rowx=0, colx=j).value: sheet.cell(rowx=i, colx=j).value})
#
#             break
#
#     # Return the dictionary
#     return getValues

def fncGetValues(loc, sheetName, testCase):
    # import openpyxl module to use excel operations
    import openpyxl
    from openpyxl import load_workbook

    # Create workbook object and capture the active sheet
    wb = load_workbook(loc, data_only=True)
    sheet = wb.active

    # Designate the actual sheetname
    sheet = wb[sheetName]

    # Leaving the header row, go through each line to find a match with test case you want to execute
    for i in range(2, sheet.max_row + 1):
        if (sheet.cell(row=i, column=1).value == testCase):

            # if the matching test case is found, create a dictionary with data in second column
            getValues = {sheet.cell(row=1, column=2).value: sheet.cell(row=i, column=2).value}

            # Once the data dictionary is added, keep on adding rest of the items in that row till the end of columns
            for j in range(3, sheet.max_column + 1):
                getValues.update({sheet.cell(row=1, column=j).value: sheet.cell(row=i, column=j).value})

            break

    # Return the dictionary
    return getValues


def fncSelectElementFromDropdown(myDriver, idDDElement, uElement):
    import time

    # Click on the dropdown to set context
    myDriver.find_element_by_id(idDDElement).click()
    time.sleep(2)

    # Starting from item 2 in the dropdown, compare the text with input value and select it
    i = 2
    found = False
    while (found == False):
        xPathdropdownItem = '//*[@id="'+idDDElement +'_DropDown"]/div/ul/li['+str(i)+']'
        if (myDriver.find_element_by_xpath(xPathdropdownItem).text == uElement):
            myDriver.find_element_by_xpath(xPathdropdownItem).click()
            found = True
        else:
            i += 1
            found = False

def fncEnterTextboxValue(myDriver, idtxtElement, uElement):
    myDriver.find_element_by_id(idtxtElement).click()
    myDriver.find_element_by_id(idtxtElement).clear()
    myDriver.find_element_by_id(idtxtElement).send_keys(str(uElement))

def fncGetFinalPremiumFromRater(uHREF):
    import xlrd
    import locale

    uRaterFile = uHREF.split('=')[1]
    uRaterFileLocal = 'C:\\Users\\niraj\\Downloads\\' + uRaterFile

    # Create workbook object and capture the active sheet
    wb = xlrd.open_workbook(uRaterFileLocal)

    # Designate the actual sheetname
    sheet = wb.sheet_by_name('XS Final Premium Calculation')

    locale.setlocale( locale.LC_ALL, 'English_United States.1252' )

    # Get Final Adjustment Premium
    uFinalPremium = sheet.cell(rowx=34, colx=1).value
    ExFinalPremium = locale.currency( uFinalPremium, grouping = True )

    # Get Surplus Lines State Tax
    # uSurplusLinesStateTax = sheet.cell(rowx=36, colx=1).value
    # ExSurplusLinesStateTax = locale.currency( uSurplusLinesStateTax, grouping = True )

    # return ExFinalPremium, ExSurplusLinesStateTax

    return ExFinalPremium

# __________________________________________________________________________
# Checking GitHub Process
# __________________________________________________________________________
